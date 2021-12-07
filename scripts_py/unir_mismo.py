import pandas as pd
from openpyxl import load_workbook
import scripts_py.unir_mismo as unir_mismo
# @title Script funciones
def unir(archivo):
    b1 = None
    excel = load_workbook(archivo)
    hojas = excel.sheetnames

    for x in hojas:
        df = pd.read_excel(archivo, x)
        frames1 = [df, b1]
        result = pd.concat(frames1, ignore_index=True, sort=False)
        ant = result

        if result is None:
            b1 = None
        else:
            b1 = ant
    return result